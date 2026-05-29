"""Convert the user-supplied 2026 Predictions Excel into web-ready JSON.

Reads `GB_2026_REVISED_Predictions_28May.xlsx` from the repo root and emits:

  data/exports/predictions_2026_revised.json
  data/exports/predictions_2026_methodology.json
  data/exports/predictions_2026_summary.json
  web/public/data/predictions_2026_revised.json        (same content, web copy)
  web/public/data/predictions_2026_methodology.json    (same)
  web/public/data/predictions_2026_summary.json        (same)

The source file is a qualitative human-analyst model, not a machine-learning
output. Confidence ranges are intentional (e.g. PTI-backed "3-4 seats"). We
preserve the prose so the dashboard can show both the verdict and the
reasoning side-by-side.

Run from anywhere via:
    uv run python -m gb_pipeline.convert_predictions
"""
from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any

import openpyxl


ROOT = Path(__file__).resolve().parents[3]
SRC_XLSX = ROOT / "GB_2026_REVISED_Predictions_28May.xlsx"
DATA_EXPORTS = ROOT / "data" / "exports"
WEB_DATA = ROOT / "web" / "public" / "data"


def _normalise_party(raw: str | None) -> dict[str, Any]:
    """Split strings like 'MWM (PTI-backed)' or 'Independent (PTI-backed)' into
    a clean party_id plus a pti_proxy flag. Returns party_id matching the keys
    in web/src/lib/parties.ts (PPP, PML-N, PTI, MWM, IPP, JUI-F, TLP,
    Independent, etc.) so the dashboard can look up flag and colour.
    """
    if not raw:
        return {"party_id": "Independent", "party_raw": "", "pti_proxy": False}
    raw = raw.strip()
    pti_proxy = "PTI-backed" in raw or "PTI proxy" in raw
    # Strip any parenthetical to get the canonical party token.
    canonical = re.sub(r"\s*\([^)]*\)\s*$", "", raw).strip()
    # Map common spellings to the IDs the web app already knows.
    mapping = {
        "PPP": "PPP",
        "PML-N": "PML-N",
        "PML N": "PML-N",
        "PMLN": "PML-N",
        "PTI": "PTI",
        "MWM": "MWM",
        "JUI-F": "JUI-F",
        "JUI F": "JUI-F",
        "JUIF": "JUI-F",
        "JUEIF": "JUI-F",
        "IPP": "IPP",
        "ITP": "ITP",
        "TLP": "TLP",
        "JI": "JI",
        "MQM": "MQM",
        "ANP": "ANP",
        "AP": "AP",
        "AWP": "AWP",
        "MML": "MML",
        "PAT": "PAT",
        "PNP": "PNP",
        "SIC": "SIC",
        "TTPP": "TTPP",
        "PPP-TOWER": "PPP-TOWER",
        "BNF-N": "BNF-N",
        "BNF": "BNF-N",
        "PML-Q": "PML-Q",
        "PML": "PML",
        "Independent": "Independent",
        "Ind": "Independent",
        "Ind.": "Independent",
    }
    party_id = mapping.get(canonical, "Independent")
    return {
        "party_id": party_id,
        "party_raw": raw,
        "pti_proxy": pti_proxy,
    }


def _parse_predicted_votes(value: Any) -> dict[str, Any]:
    """Extract a numeric estimate from a string like '~11,200' so the UI can
    render bar lengths proportionally without losing the original wording."""
    if value is None:
        return {"predicted_votes_text": "", "predicted_votes_estimate": None}
    text = str(value).strip()
    if not text:
        return {"predicted_votes_text": "", "predicted_votes_estimate": None}
    cleaned = re.sub(r"[^0-9]", "", text)
    estimate = int(cleaned) if cleaned else None
    return {"predicted_votes_text": text, "predicted_votes_estimate": estimate}


def convert() -> None:
    if not SRC_XLSX.exists():
        raise FileNotFoundError(f"Source Excel not found at {SRC_XLSX}")
    wb = openpyxl.load_workbook(SRC_XLSX, data_only=True)

    # ---- Sheet 1: per-seat predictions (72 rows = 24 seats x 3 ranks) -------
    ws = wb["REVISED Predictions"]
    headers = [c.value for c in ws[1]]
    expected = [
        "Constituency",
        "Area",
        "Rank",
        "Predicted Position",
        "Party",
        "Predicted Votes",
        "Margin",
        "Social Media Sentiment",
        "Hidden Truth / Ground Reality",
    ]
    if headers != expected:
        raise ValueError(
            f"Unexpected REVISED Predictions header: {headers!r}\n"
            f"Expected:                                {expected!r}"
        )

    rows: list[dict[str, Any]] = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        (
            cz,
            area,
            rank,
            candidate,
            party,
            votes,
            margin,
            sm_note,
            ground_note,
        ) = raw
        if not cz:
            continue
        party_info = _normalise_party(party)
        votes_info = _parse_predicted_votes(votes)
        rows.append(
            {
                "constituency_id": cz,
                "area_name": (area or "").strip(),
                "rank": int(rank) if rank is not None else None,
                "candidate_name": (candidate or "").strip(),
                **party_info,
                **votes_info,
                "margin": (margin or "").strip(),
                "social_media_sentiment": (sm_note or "").strip(),
                "ground_reality": (ground_note or "").strip(),
            }
        )

    # ---- Sheet 2: REVISED Overall Summary -----------------------------------
    ws = wb["REVISED Overall Summary"]
    summary_rows = [list(r) for r in ws.iter_rows(values_only=True)]
    # Locate the "Party/Bloc" table by scanning for that header.
    party_table: list[dict[str, Any]] = []
    flips: list[dict[str, str]] = []
    scenarios: list[dict[str, str]] = []
    mode = None
    for r in summary_rows:
        head = (r[0] or "").strip() if r else ""
        if head == "Party/Bloc":
            mode = "party"
            continue
        if head == "Constituency":
            mode = "flips"
            continue
        if head and "GOVERNMENT FORMATION SCENARIOS" in head:
            mode = "scenarios"
            continue
        if not head and mode in ("party", "flips", "scenarios") and not any(r):
            mode = None
            continue
        if mode == "party":
            if head and head not in {"Party/Bloc"}:
                party_table.append(
                    {
                        "party_or_bloc": head,
                        "seats": (str(r[1]).strip() if r[1] is not None else ""),
                        "driver": (str(r[2]).strip() if r[2] is not None else ""),
                    }
                )
        elif mode == "flips":
            if head and head != "Constituency":
                flips.append(
                    {
                        "constituency": head,
                        "flip": (str(r[1]).strip() if r[1] is not None else ""),
                        "reason": (str(r[2]).strip() if r[2] is not None else ""),
                    }
                )
        elif mode == "scenarios":
            if head and head not in {"Most Likely (65%)", "Alternative (25%)", "Chaos (10%)"} == False:
                # Hack: the scenarios rows have the label as the first col and
                # the description in the second. Detect by checking col 1.
                pass
            if head and r[1]:
                scenarios.append({"label": head, "description": str(r[1]).strip()})

    # Pull the headline strings from row 0..2.
    title_block = []
    for r in summary_rows[:4]:
        if r and r[0]:
            title_block.append(str(r[0]).strip())

    summary = {
        "title_lines": title_block,
        "party_projection": party_table,
        "critical_flips": flips,
        "government_formation_scenarios": scenarios,
        "election_date": "2026-06-07",
        "gba24_delay_note": "GBA-24 (Diamer-V) polling delayed to 15 November 2026.",
    }

    # ---- Sheet 4: Methodology blob ------------------------------------------
    ws = wb["Methodology & Sources"]
    methodology_blob = ""
    for r in ws.iter_rows(values_only=True):
        for cell in r:
            if cell:
                methodology_blob = str(cell)
                break
        if methodology_blob:
            break

    methodology = {
        "title": "Revised methodology for GB 2026 constituency predictions",
        "revision": "2.0",
        "prediction_date": "2026-05-28",
        "full_text": methodology_blob.strip(),
    }

    # ---- Write outputs ------------------------------------------------------
    DATA_EXPORTS.mkdir(parents=True, exist_ok=True)
    WEB_DATA.mkdir(parents=True, exist_ok=True)
    for target_dir in (DATA_EXPORTS, WEB_DATA):
        (target_dir / "predictions_2026_revised.json").write_text(
            json.dumps(rows, indent=2, ensure_ascii=False) + "\n", encoding="utf-8"
        )
        (target_dir / "predictions_2026_summary.json").write_text(
            json.dumps(summary, indent=2, ensure_ascii=False) + "\n", encoding="utf-8"
        )
        (target_dir / "predictions_2026_methodology.json").write_text(
            json.dumps(methodology, indent=2, ensure_ascii=False) + "\n", encoding="utf-8"
        )

    print(f"Wrote {len(rows)} predictions across "
          f"{len({r['constituency_id'] for r in rows})} seats")
    print(f"Party projection rows: {len(party_table)}")
    print(f"Critical flips: {len(flips)}")
    print(f"Government scenarios: {len(scenarios)}")
    print(f"Methodology chars: {len(methodology_blob)}")


if __name__ == "__main__":
    convert()
